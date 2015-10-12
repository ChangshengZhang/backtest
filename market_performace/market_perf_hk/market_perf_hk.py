# -*- coding: utf-8 -*-
# File Name: market_perf.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

#########################################################################

from WindPy import * 
from datetime import * 
import copy
from openpyxl import load_workbook
from openpyxl import Workbook
import time

class MarketPerfHk():
	"""docstring for get_market_close_data"""
	def __init__(self):
		if w.isconnected() == False:
			w.start()

		self.close_data = []
		self.stock_name = []
		self.susp_stock_name = []
		self.rt_chg = []
		self.close_info = []
		self.all_stock_num = 0
		self.susp_stock_num = 0
		self.stock_mark_up = []
		self.stock_mark_down = []

		
		self.run()
		

	def get_init_data(self):
		file_path = 'input/market_perf_hk.xlsx'
		wb = load_workbook(file_path)
		sheet_names = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheet_names[0])
		self.stock_name =[]

		row_num = len(ws.rows)
		for ii in range(1,row_num):
			temp = ws.cell(row=ii+1,column=2).value
			a = temp.split(" ")
			code_name = ""
			for jj in range(5-len(a[0])):
				code_name = code_name+"0"
			code_name = code_name+a[0]+"."+"hk"
			self.stock_name.append(code_name)

		self.all_stock_num = len(self.stock_name)


	def get_market_close_data(self):

		type_a = 0
		type_b = 0 
		type_c = 0 
		type_d = 0
		type_e = 0 
		type_f = 0
		# 0% or suspend
		type_g = 0
		self.stock_mark_down = []
		self.close_info = []
		self.rt_chg = []
		self.stock_mark_up = []

		res = w.wsq(self.stock_name,"rt_pct_chg")

		if res.ErrorCode != 0:
			print('Error['+str(res.ErrorCode)+'][load stockcode list fail]\n')
			print "rt_pct_chg error"
			sys.exit()

		if res.ErrorCode == 0:
			for item in res.Data[0]:
				self.rt_chg.append(float(item))

		for ii in range(len(self.rt_chg)):
			if self.rt_chg[ii]>=0.099:
				type_a = type_a +1 
				self.stock_mark_up.append(self.stock_name[ii])

			elif self.rt_chg[ii]>=0.05:
				type_b = type_b +1
				
			elif self.rt_chg[ii] >0:
				type_c = type_c +1 
			elif self.rt_chg[ii] ==0:
				type_g = type_g +1
			elif self.rt_chg[ii] >-0.05:
				type_d = type_d + 1 
			elif self.rt_chg[ii] > -0.099:
				type_e = type_e +1 
			else:
				type_f = type_f +1 
				self.stock_mark_down.append(self.stock_name[ii])
		#time_now = str(datetime.today()).split(".")[0].split(" ")
		time_str = str(datetime.today()).split(".")[0].split(" ")
		temp_str = time_str[0].split("-")
		temp_str_1 = time_str[1].split(":")
		time_now = temp_str[0]+temp_str[1]+temp_str[2] + temp_str_1[0]+temp_str_1[1]


		self.close_info =[time_now,self.all_stock_num,type_a,type_b,type_c,type_g,type_d,type_e,type_f]
		

	def write_to_excel(self):
		file_path = "data/data_hk.xlsx"
		wb = load_workbook(file_path)
		sheet_names = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheet_names[0])

		temp_row = len(ws.rows)+1
		for ii in range(len(ws.columns)):

			ws.cell(row =temp_row ,column =ii+1).value = self.close_info[ii]

		wb.save(file_path)


	def run(self):
		self.get_init_data()

		for ii in range(60):
			self.get_market_close_data()
			self.write_to_excel()
			print "the "+str(ii) + " th is done.\n"
			time.sleep(600)
		#self.fix_data_excel()


if __name__ == '__main__':

	market_close_data = MarketPerfHk()



