# -*- coding: utf-8 -*-
# File Name: main.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

#########################################################################

import os
import sys
import strategy.HighLowPoint as HighLowPoint
import data.LoadUserData as LoadUserData
import data.LoadData as LoadData
import time
import datetime
import output.SendEmail as SendEmail
from openpyxl import load_workbook

class HighLowPointClass():
	"""docstring for HighLowPointClass"""
	def __init__(self):
		file_path = "input/HighLowPoint/user_data.xlsx"

		stock_code_list,stock_name_list,stock_buy_price_list,stock_sell_price_list = LoadUserData.load_user_data(file_path)
		stock_daily_data_list = LoadData.get_daily_stock_data(stock_code_list)
		self.fall_end_flag = []
		self.rise_end_flag = []

		for ii in range(len(stock_code_list)):
			self.fall_end_flag.append(0)
			self.rise_end_flag.append(0)

		print "begin running\n"
		self.run(stock_code_list,stock_buy_price_list,stock_sell_price_list)


	def run(self,stock_code_list,stock_buy_price_list,stock_sell_price_list):

		for ii in range(60):
			print "the " + str(ii) + " th is begining"
			self.get_high_low_points(stock_code_list)
			#self.stock_price_reminder(stock_code_list,stock_buy_price_list,stock_sell_price_list)
			print "the "+ str(ii) + " th is done."
			time.sleep(300)


	def stock_price_reminder(self,stock_code_list,stock_buy_price_list,stock_sell_price_list):
		
		realtime_price_list = LoadData.get_realtime_price(stock_code_list)
		reminder_info = []

		for ii in range(len(stock_code_list)):

			if float(realtime_price_list[ii]) >=float(stock_sell_price_list[ii]):
				#sell
				temp_reminder_info = str(stock_code_list[ii])+" should be sold, whose target sell price is "+str(stock_sell_price_list[ii])+", now price is "+str(realtime_price_list[ii])
				reminder_info.append(temp_reminder_info)
			elif float(realtime_price_list[ii])<=float(stock_buy_price_list[ii]):
				
				temp_reminder_info = str(stock_code_list[ii])+" should be bought, whose target buy price is "+str(stock_buy_price_list[ii])+ ", now price is "+str(realtime_price_list[ii])
				reminder_info.append(temp_reminder_info)

		if len(reminder_info)!= 0:
			mail_content = ""
			for jj in range(len(reminder_info)):
				mail_content = mail_content + "\n" + reminder_info[jj]+"\n"
			print mail_content 
			b = SendEmail.Send_Email("Stock price reminder from H&L Model",mail_content)


	def get_high_low_points(self,stock_code_list):
		msg = '''<html> <tr>Trend inverse reminder from H&L Model</tr> <table width="300" border="1" bordercolor="black" cellspacing="1">'''
		bar_size = [5,30,60]
		column_num = [5,9,13]
		threshold = [0.03,0.05,0.15]
		send_email_flag = False
		for ii in range(len(bar_size)):
			if bar_size[ii]==60:
				stock_intraday_data = LoadData.get_daily_stock_data(stock_code_list,"20100101")
				msg = msg +''' <tr><td rowspan="2">'''  + ''' daily </td>'''
			else:
				stock_intraday_data = LoadData.get_intraday_stock_data(stock_code_list,bar_size=bar_size[ii],delta_days = 365)
				msg = msg +''' <tr><td rowspan="2">''' +str(bar_size[ii]) + ''' min </td>'''
			high_point_index_list,high_point_list,high_point_time_list,low_point_index_list,low_point_list,low_point_time_list = HighLowPoint.get_high_low_points_list(stock_intraday_data,threshold[ii])
			
			
			fall_end_list = []
			rise_end_lsit = []
			fall_end_msg = ""
			rise_end_msg = ""
			for jj in range(len(stock_code_list)):

				self.write_data_to_file(column_num[ii],jj,high_point_list[jj][-1],high_point_time_list[jj][-1],low_point_list[jj][-1],low_point_time_list[jj][-1])
				
				temp_msg_fall_end, temp_msg_rise_end = self.judge_trend_inverse(high_point_list[jj],low_point_list[jj],jj) 

				if temp_msg_fall_end =="True":
					fall_end_list.append(stock_code_list[jj])
					send_email_flag = True
				if temp_msg_rise_end == "True":
					rise_end_lsit.append(stock_code_list[jj])
					send_email_flag = True
			for item in fall_end_list:
				fall_end_msg = fall_end_msg + item + " "
			for item in rise_end_lsit:
				rise_end_msg = rise_end_msg + item +" "

			msg = msg +"<td>" + "Falling End" + "</td><td>"+fall_end_msg +"</td></tr><tr><td> Rising End </td><td>"+rise_end_msg+"</td></tr> \n"
		msg = msg +" </table></html>"
		print msg
		if send_email_flag == True:
			a = SendEmail.Send_Email("Trend inverse reminder from H&L Model",msg)
			print a.isSend
		print self.fall_end_flag
		print self.rise_end_flag


	def judge_trend_inverse(self,high_point_list,low_point_list,current_price):
		number = min(len(high_point_list),len(low_point_list))
		high_list_len = len(high_point_list)
		low_list_len = len(low_point_list)
		action_type_list = []
		action_index_list = []
		pre_high_flag = 0
		pre_low_flag = 0
		pos_status = 0

		down_flag = False
		up_flag = False

		for ii in range(number):
			long_flag = 0 
			short_flag = 0
			
			if float(high_point_list[ii+high_list_len-number])<float(ii+high_list_len-number):
				pre_high_flag = pre_high_flag +1

			else:
				if pre_high_flag >=1 and(len(action_type_list)==0 or action_type_list[-1]!="l"):
					long_flag =1
				pre_high_flag = 0 
			if float(low_point_list[ii+low_list_len-number])>float(ii+low_list_len-number):
				pre_low_flag = pre_low_flag +1
			else:
				if pre_low_flag >=1 and(len(action_type_list)==0 or action_type_list[-1]!="s"):
					short_flag =1 
				pre_low_flag = 0

			if pos_status !=1 and long_flag ==1:
				action_type_list.append("l")
				action_index_list.append(ii)
				pos_status = 1
			if pos_status !=-1 and short_flag ==1:
				action_type_list.append("s")
				action_index_list.append(ii)
				pos_status = -1
		
		if len(action_type_list)!=0:
			if action_type_list[-1] == "l":
				if current_price < low_point_list[-1]:
					down_flag = True
			else:
				if current_price > high_point_list[-1]:
					up_flag = True

		return up_flag,down_flag





	# def judge_trend_inverse(self,high_point_list,low_point_list,jj):

	# 	msg_fall_end = "False"
	# 	msg_rise_end = "False"
	# 	if len(high_point_list)>=3:
	# 		if high_point_list[-1]> high_point_list[-3] and high_point_list[-2]<high_point_list[-3]:
	# 			if self.fall_end_flag[jj]==0:
	# 				msg_fall_end = "True"
	# 				self.fall_end_flag[jj]=1
	# 		else:
	# 			self.fall_end_flag[jj]=0

	# 	if len(low_point_list)>=3 :
	# 		if low_point_list[-1]<low_point_list[-3] and low_point_list[-2] > low_point_list[-3]:
	# 			if self.rise_end_flag[jj]==0:
	# 				msg_rise_end = "True"
	# 				self.rise_end_flag[jj]=1
	# 		else:
	# 			self.rise_end_flag[jj]=0

	# 	return msg_fall_end,msg_rise_end 

	def write_data_to_file(self,column_num,jj,high_point,high_point_time,low_point,low_point_time):
		file_name = "input/HighLowPoint/user_data.xlsx"
		wb = load_workbook(file_name)
		sheet_names = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheet_names[0])

		
		ws.cell(row=jj+2,column = column_num).value = str(high_point)
		ws.cell(row=jj+2,column = column_num+1).value = str(high_point_time)
		ws.cell(row=jj+2,column = column_num +2).value = str(low_point)
		ws.cell(row = jj+2,column = column_num+3).value = str(low_point_time)

		wb.save(file_name)



if __name__ == '__main__':
	
	a = HighLowPointClass()