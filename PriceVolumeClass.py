# -*- coding: utf-8 -*-
# File Name: pricevolume.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

#########################################################################

import strategy.PriceVolume as PriceVolume
import data.LoadData as LoadData
import output.PlotChart as PlotChart
import matplotlib.pyplot as plt
import output.SendEmail as SendEmail
from openpyxl import load_workbook
from datetime import *

class PriceVolumeClass():
	"""docstring for PriceVolumeClass"""
	def __init__(self):
		input_file_path = "input/PriceVolume/bench_mark.xlsx"

		bench_mark_list,self.bench_mark_name_list = self.load_bench_mark_data(input_file_path)

		#bench_mark_list = ["000016.SH","000001.SH","000300.SH","399006.SZ","399905.SZ"]
		#self.bench_mark_name_list = [u"上证50",u"上证综指",u"沪深300",u"创业板指",u"中证500"]
		#bench_mark_list = ["0806.HK","QIHU.N"]
		#self.bench_mark_name_list = [u"惠理",u"奇虎"]
		stock_data_list = LoadData.get_daily_stock_data(bench_mark_list,"20110101")
		self.price_index = 4
		self.volume_index = 5 

		self.trade_cost = 1-0.001
		self.range_day = 1
		self.regression_day = [2,5]
		
		# short flag = 1 表示可以做空
		short_flag = 1
		
		self.price_list = []
		self.volume_list = []

		for ii in range(len(bench_mark_list)):
			self.price_list.append(list(zip(*stock_data_list[ii])[self.price_index]))
			self.volume_list.append(list(zip(*stock_data_list[ii])[self.volume_index]))

		self.plot(self.price_list,bench_mark_list,self.range_day,short_flag)

		#self.run(bench_mark_list)

		print "done\n"
		
	def load_bench_mark_data(self,input_file_path):
		wb = load_workbook(input_file_path)
		sheet_names = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheet_names[0])
		bench_mark_list = []
		bench_mark_name_list = []
		row_num = len(ws.rows)
		for ii in range(2,row_num):
			temp_bench_mark = str(ws.cell(row=ii+1,column = 1).value)
			temp_bench_mark_name = ws.cell(row= ii+1,column =2).value
			bench_mark_list.append(temp_bench_mark)
			bench_mark_name_list.append(temp_bench_mark_name)

		return bench_mark_list,bench_mark_name_list



	def identity(self,price_list,revenue_list,action_index_list):

		new_revenue_list =[]
		new_action_index = []
		for ii in range(len(price_list)-len(revenue_list)):
			new_revenue_list.append(1.0)
		for ii in range(len(revenue_list)):
			new_revenue_list.append(revenue_list[ii])
		for ii in range(len(action_index_list)):
			new_action_index.append(action_index_list[ii]+len(price_list)-len(revenue_list))
		return new_revenue_list,new_action_index

	def plot(self,price_list,bench_mark_list,range_day,short_flag=1):
		color =["b","y","c","k"]
			
		cc =0
		for ii in range(len(self.regression_day)):
			
			self.offset = range_day +self.regression_day[ii] -2
			self.slope_price_list = PriceVolume.PriceVolume(self.price_list,range_day,self.regression_day[ii])
			self.slope_volume_list = PriceVolume.PriceVolume(self.volume_list,range_day,self.regression_day[ii])

			action_index_list,action_type_list,revenue_list = self.backtest(self.offset,short_flag)
			for jj in range(len(bench_mark_list)):
				plt.figure(jj)
				print bench_mark_list[jj],"regression_day:",self.regression_day[ii],len(action_index_list[jj]),len(self.price_list[jj])
				new_revenue, new_action_index = self.identity(self.price_list[jj],revenue_list[jj],action_index_list[jj])

				temp_long_action_index = []
				temp_short_action_index = []
				for kk in range(len(new_action_index)):
					if action_type_list[jj][kk] == "b":
						temp_long_action_index.append([new_action_index[kk],new_revenue[new_action_index[kk]]])
					else:
						temp_short_action_index.append([new_action_index[kk],new_revenue[new_action_index[kk]]])
				volume_index = []
				volume_num = []
				max_volumn = 2.0*max(self.volume_list[jj])
				for kk in range(len(self.volume_list[jj])):
					volume_index.append(kk)
					volume_num.append(self.volume_list[jj][kk]/max_volumn)

				plt.plot(new_revenue,color = color[cc%len(color)],linewidth = 2.5,label = str(self.regression_day[ii])+u"天回归的策略收益")
				
				plt.scatter(list(zip(*temp_long_action_index)[0]),list(zip(*temp_long_action_index)[1]),c="r",s=70,marker="s")
				plt.scatter(list(zip(*temp_short_action_index)[0]),list(zip(*temp_short_action_index)[1]),c="g",s=70,marker="v")
			cc = cc +1

		for ii in range(len(bench_mark_list)):
			plt.figure(ii)
			plt.plot(self.norm_list(price_list[ii]),color = "r",label =u"价格走势")
			plt.bar(volume_index,volume_num,color ="k",label= u"成交量")
			plt.legend(loc="upper right")
			plt.xlim(0,len(revenue_list[0])+30)
			plt.ylim(0)
			if short_flag ==1:
				plt.title("long/short "+self.bench_mark_name_list[ii])
			else:
				plt.title("long/hold "+self.bench_mark_name_list[ii])
			plt.xlabel("Time /day")
			plt.ylabel("Return")
			plt.grid(True)
			fig = plt.gcf()
			fig.set_size_inches(24, 15)
			fig.savefig("png/PriceVolume/"+bench_mark_list[ii]+".png",dpi = 200)

		#plt.show()

	def norm_list(self,list_a):
		new_list =[]
		for ii in range(len(list_a)):
			new_list.append(list_a[ii]/list_a[0])
		return new_list
	#short_flag =1表示能short
	def backtest(self,offset,short_flag=0):
		action_index_list = []
		action_type_list =[] 
		revenue_list = []

		for ii in range(len(self.slope_price_list)):
			pos_flag = 0
			temp_action_index_list = []
			temp_action_type_list = []
			temp_revenue_list = []
			
			for jj in range(len(self.slope_price_list[ii])):
				if self.slope_price_list[ii][jj] ==1 and self.slope_volume_list[ii][jj]== 1:
					if pos_flag ==0:
						if len(temp_action_index_list)==0:
							temp_revenue_list.append(1.0*self.trade_cost)
						else:
							temp_revenue_list.append(temp_revenue_list[-1]*self.trade_cost)

						
						temp_action_type_list.append("b")
						temp_action_index_list.append(jj)
						pos_flag =1
					else:
						temp_revenue_list.append(temp_revenue_list[-1]*float(self.price_list[ii][jj+offset])/float(self.price_list[ii][jj+offset-1]))

				else:
					if pos_flag ==1:

						if self.slope_price_list[ii][jj]==-1 and self.slope_volume_list[ii][jj] ==-1:
							temp_revenue_list.append(temp_revenue_list[-1]*float(self.price_list[ii][jj+offset])/float(self.price_list[ii][jj+offset-1])*self.trade_cost)
							temp_action_index_list.append(jj)
							temp_action_type_list.append("s")
							pos_flag = 0
						else:
							temp_revenue_list.append(temp_revenue_list[-1]*float(self.price_list[ii][jj+offset])/float(self.price_list[ii][jj+offset-1]))
					else:
						if len(temp_action_index_list)==0:
							temp_revenue_list.append(1.0)
						else:
							if short_flag ==0:
								temp_revenue_list.append(temp_revenue_list[-1])
							else:
								temp_revenue_list.append(temp_revenue_list[-1]*(2-float(self.price_list[ii][jj+offset]/float(self.price_list[ii][jj+offset-1]))))

			revenue_list.append(temp_revenue_list)
			action_index_list.append(temp_action_index_list)
			action_type_list.append(temp_action_type_list)
		return action_index_list,action_type_list,revenue_list


	def run(self,bench_mark_list,short_flag = 1,range_day =1):
		realtime_price_list,realtime_volume_list = LoadData.get_realtime_price_and_volume(bench_mark_list)
		print bench_mark_list
		print realtime_price_list
		print realtime_volume_list

		reminder_info_flag_list = []
		is_send_email = False


		for ii in range(len(bench_mark_list)):

			if float(datetime.today().hour) >=15:
				self.price_list[ii].pop()
				self.volume_list[ii].pop()
			print "volume",self.volume_list[ii][-1]

		old_action_len = []


		for regression_day in self.regression_day:

			old_action_len_per_rangeday = []
			self.offset = range_day +regression_day -2
			self.slope_price_list = PriceVolume.PriceVolume(self.price_list,range_day,regression_day)
			self.slope_volume_list = PriceVolume.PriceVolume(self.volume_list,range_day,regression_day)
			action_index_list,action_type_list,revenue_list = self.backtest(self.offset,short_flag)

			for ii in range(len(action_type_list)):
				old_action_len_per_rangeday.append(len(action_index_list[ii]))
			old_action_len.append(old_action_len_per_rangeday)

		for ii in range(len(bench_mark_list)):
			self.price_list[ii].append(realtime_price_list[ii])
			self.volume_list[ii].append(realtime_volume_list[ii])
			print "volume new",self.volume_list[ii][-1]

		for ii in range(len(self.regression_day)):

			temp_reminder_info_flag = []
			regression_day = self.regression_day[ii]
			

			self.offset = range_day +regression_day -2
			self.slope_price_list = PriceVolume.PriceVolume(self.price_list,range_day,regression_day)
			self.slope_volume_list = PriceVolume.PriceVolume(self.volume_list,range_day,regression_day)
			action_index_list,action_type_list,revenue_list = self.backtest(self.offset,short_flag)

			for jj in range(len(action_index_list)):
				print len(action_index_list[jj]),old_action_len[ii][jj]
				if len(action_index_list[jj]) - old_action_len[ii][jj] ==1:
					temp_reminder_info_flag.append(action_type_list[jj][-1])
					is_send_email = True
				else:
					temp_reminder_info_flag.append("x")

			reminder_info_flag_list.append(temp_reminder_info_flag)

		print reminder_info_flag_list

		if is_send_email == True:
			msg= '''<html> <tr>Trade  reminder from Price & Volume Model</tr> <table width="300" border="1" bordercolor="black" cellspacing="1">'''

			for ii in range(len(reminder_info_flag_list)):
				if ii ==0:
					msg = msg + '''<tr><td> ''' + '''   </td>'''
					for jj in range(len(reminder_info_flag_list[ii])):
						msg = msg + ''' <td>  ''' + str(bench_mark_list[jj])+ "</td>"
					msg = msg +  '''  </tr>\n'''
				msg = msg + '''<tr><td>''' + str(self.regression_day[ii]) + ''' day </td>'''
				for jj in range(len(reminder_info_flag_list[ii])):
					if reminder_info_flag_list[ii][jj] =="b":
						msg = msg + '''<td> long </td>'''
					elif reminder_info_flag_list[ii][jj] =="s":
						if short_flag ==1:
							msg = msg + '''<td> short </td> '''
						else:
							msg = msg + '''<td> close previous long postion </td>'''
					else:
						msg = msg + '''<td> </td> '''

				msg = msg + '''</tr> \n'''
			msg = msg + '''</table></html> '''

			send_email = SendEmail.Send_Email("Trade  reminder from Price & Volume Model",msg)
			print send_email.isSend



if __name__ == '__main__':
	a = PriceVolumeClass()




