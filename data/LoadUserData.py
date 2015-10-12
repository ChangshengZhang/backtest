# -*- coding: utf-8 -*-
# File Name: LoadUserData.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

##############################################

from openpyxl import load_workbook
from openpyxl import Workbook

def load_user_data(file_path):

	wb = load_workbook(file_path)
	sheet_names = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheet_names[0])

	user_stock_code_list = []
	user_stock_name_list = []
	user_stock_buy_point_list = []
	user_stock_sell_point_list = []
	
	row_num = len(ws.rows)
	for ii in range(1,row_num):
		user_stock_code_list.append(ws.cell(row=ii+1,column=1).value)
		user_stock_name_list.append(ws.cell(row=ii+1,column=2).value)
		user_stock_buy_point_list.append(float(ws.cell(row=ii+1,column=3).value))
		user_stock_sell_point_list.append(float(ws.cell(row=ii+1,column =4).value))

	return user_stock_code_list,user_stock_name_list,user_stock_buy_point_list,user_stock_sell_point_list


def load_user_stock_name_list(file_path):

	wb = load_workbook(file_path)
	sheet_names = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheet_names[0])

	user_stock_code_list = []
	row_num = len(ws.rows)
	for ii in range(1,row_num):
		user_stock_code_list.append(ws.cell(row=ii+1,column=1).value)

	return user_stock_code_list

def load_user_market_perf_hk_list(file_path):
	wb = load_workbook(file_path)
	sheet_names = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheet_names[0])
	user_marktet_perf_hk_list =[]

	row_num = len(ws.rows)
	for ii in range(1,row_num):
		temp = ws.cell(row=ii+1,column=1).value
		a = temp.split(" ")
		code_name = ""
		for jj in range(5-len(a[0])):
			code_name = code_name+"0"
		code_name = code_name+a[0]+"."+"hk"
		user_marktet_perf_hk_list.append(code_name)
	return user_marktet_perf_hk_list

