# -*- coding: utf-8 -*-
# File Name: DeviationFromMA .py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com
# Created Time: Tue Jul  7 10:09:08 2015

#########################################################################

import sys
sys.path.append("..")
import scipy.stats as stats
import data.CalcIndex as CalcIndex
import data.LoadData as LoadData
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

class DeviationFromMA():
	"""docstring for DeviationFromMA"""
	def __init__(self,ma_days=200,start_date="19950101",init_days = 60,short_flag =0):
		
		input_file_path = "../input/DeviationFromMA/user_data.xlsx"
		self.user_stock_code_list,self.user_stock_name_list = self.load_excel(input_file_path)
		self.user_stock_data_list = LoadData.get_daily_stock_data(self.user_stock_code_list,start_date)
		self.short_flag = short_flag
		self.trade_cost = 1-0.001
		self.ma_days = ma_days
		self.user_stock_price_data_list = []
		self.user_stock_uniform_price_list =[]
		self.user_stock_datetime_list = []
		self.user_stock_price_data_old_list =[]
		for ii in range(len(self.user_stock_code_list)):
			self.user_stock_price_data_list.append((list(zip(*self.user_stock_data_list[ii])[4]))[init_days:])
			self.user_stock_price_data_old_list.append(list(zip(*self.user_stock_data_list[ii])[4]))
			self.user_stock_datetime_list.append((list(zip(*self.user_stock_data_list[ii])[0]))[init_days:])
			self.user_stock_uniform_price_list.append(CalcIndex.Uniform_Price(self.user_stock_price_data_list[ii]))

		self.price_ma_list = CalcIndex.Calc_MA(self.user_stock_price_data_old_list,int(ma_days),int(init_days))

		self.diviation_from_ma_list,self.diviation_from_ma_std_list = self.calc_diviation_from_ma(ma_days)

		self.ma_dict ={60:10,120:20,200:60,250:60}
		
		self.diviation_ma_list = CalcIndex.Calc_MA(self.diviation_from_ma_list,self.ma_dict[self.ma_days])

		#self.plot(ma_days)
		#self.eveluate_index(ma_days)

		long_stand_deviation = {60:-0.5,120:-0.1,200:-0.15,250:-0.2}
		short_stand_deviation = {60:0.5,120:0.1,200:0.15,250:0.2}

		self.back_test_info = self.back_test(long_stand_deviation[ma_days],short_stand_deviation[ma_days])
		self.plot_back_test(ma_days)

		print "done."

	def plot_back_test(self,ma_days):
		for ii in range(len(self.user_stock_code_list)):
			plt.figure(ii)
			ax = plt.gca()

			std =[]
			negetive_std = []
			for jj in range(len(self.diviation_from_ma_list[ii])):
				temp_std = np.std(self.diviation_from_ma_list[ii][max(jj-ma_days+1,0):jj+1])
				std.append(temp_std)
				negetive_std.append(-1.0*temp_std)
			
			plt.plot(self.user_stock_uniform_price_list[ii],color = "k",linewidth = 2.5,label = "Price")
			#plt.plot(self.back_test_info[5][ii],color = "b",linewidth =2.5,label = "Return")
			plt.scatter(self.back_test_info[1][ii],self.back_test_info[2][ii],c="r",s=150,marker ="s",label ="Long")
			plt.scatter(self.back_test_info[3][ii],self.back_test_info[4][ii],c="g",s=150,marker ="s",label = "Short/Sell")
			plt.xlabel("Time /day")
			plt.ylabel("Return")
			plt.legend(loc="upper left")

			ax_2 = plt.twinx()
			ax_2.plot(self.diviation_from_ma_list[ii],color = "y",linewidth = 2.5,label = "Diviation From MA")
			#ax_2.plot(self.diviation_from_ma_std_list[ii],color = "r",linewidth = 2.5,label= "Positive Std")
			#ax_2.plot(negetive_std,color = "g",linewidth = 2.5,label = "Negative Std")

			#for jj in range(len(self.ma_days_list)):
			ax_2.plot(self.diviation_ma_list[ii],color ="c",linewidth = 2.5, label =str(self.ma_dict[self.ma_days])+" MA of Diviation")

			ax_2.legend(loc = "upper right")
			ax_2.set_ylabel("Diviation From MA")
			
			#plt.ylim(0)
			plt.xlim(0)

			xticks = ax.get_xticks()
			offset = xticks[1]-xticks[0]
			date_time = []
			for jj in range(len(xticks)):
				if jj *int(offset)>= len(self.user_stock_datetime_list[ii]):
					break
				date_time.append(self.user_stock_datetime_list[ii][jj*int(offset)])
			ax.set_xticklabels(date_time,rotation = 30)

			if self.short_flag ==1:
				plt.title(self.user_stock_name_list[ii]+"long/short return "+str(ma_days)+" MA")
			else:
				plt.title(self.user_stock_name_list[ii]+"long/hold return "+str(ma_days)+" MA")
			plt.grid(True)
			fig = plt.gcf()
			fig.set_size_inches(24, 15)
			fig.savefig("../input/DeviationFromMA/png_"+str(ma_days)+"/"+self.user_stock_code_list[ii]+"_return.png",dpi = 100)
			plt.close(fig)

	def buy_signal(self,ii,jj,stand_deviation):
		flag = False
		buy_signal_1 = False
		buy_signal_2 = False
		buy_signal_3 = False
		coeff = {60:14,120:7,200:6,250:5}
		coeff_1 = {60:1.5,120:2.0,200:2.5,250:2.5}
		coeff_2 = {60:8,120:4,200:2,250:1}
		if self.diviation_ma_list[ii][jj]>self.diviation_ma_list[ii][max(jj-1,0)] and self.diviation_ma_list[ii][jj]<1.0*np.percentile(self.diviation_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj+1],2)and self.diviation_ma_list[ii][jj] >min(self.diviation_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj+1])and(max(self.diviation_ma_list[ii][max(0,jj-2):jj])<self.diviation_ma_list[ii][jj]) :
			buy_signal_1 = True
		# if jj >0 and self.diviation_from_ma_list[ii][jj] <stand_deviation and self.diviation_from_ma_list[ii][jj-1]> stand_deviation:
		# 	buy_signal_1 = True

		if self.diviation_ma_list[ii][jj] < -1.1*coeff_1[self.ma_days]*self.diviation_from_ma_std_list[ii][jj] and self.diviation_ma_list[ii][jj]<stand_deviation :
			buy_signal_2 = True

		if jj >0 and self.diviation_ma_list[ii][jj-1] == min(self.diviation_ma_list[ii][max(jj-coeff_2[self.ma_days]*self.ma_days+1,0):jj]) and self.diviation_ma_list[ii][jj]>self.diviation_ma_list[ii][jj-1]:
			buy_signal_3 = True

		if buy_signal_1 or buy_signal_2 or buy_signal_3:
			flag = True

		return flag

	def sell_signal(self,ii,jj,stand_deviation):
		flag = False
		coeff = {60:14,120:7,200:6,250:5}
		coeff_1 = {60:1.5,120:2,200:2.7,250:2.7}
		coeff_2 = {60:8,120:4,200:2,250:1}
		sell_signal_1 = False
		sell_signal_2 = False
		sell_signal_3 = False
		if self.diviation_ma_list[ii][jj] <self.diviation_ma_list[ii][max(jj-1,0)] and self.diviation_ma_list[ii][jj]> 1.5*np.percentile(self.diviation_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj+1],98) and self.diviation_ma_list[ii][jj] <max(self.diviation_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj+1]) and (min(self.diviation_ma_list[ii][max(0,jj-2):jj])>self.diviation_ma_list[ii][jj]):
			sell_signal_1 = True

		#if self.diviation_from_ma_list[ii][jj] >= max(self.diviation_from_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj+1]) and self.diviation_from_ma_list[ii][jj] >= 1.1*max(self.diviation_from_ma_list[ii][max(jj-coeff[self.ma_days]*self.ma_days+1,0):jj]) :
			#sell_signal_2 = True

		if self.diviation_ma_list[ii][jj] >1.3*coeff_1[self.ma_days]*self.diviation_from_ma_std_list[ii][jj] and self.diviation_ma_list[ii][jj] > stand_deviation:
			sell_signal_2 = True

		if jj >0 and self.diviation_ma_list[ii][jj-1] == max(self.diviation_ma_list[ii][max(jj-coeff_2[self.ma_days]*self.ma_days+1,0):jj]) and self.diviation_ma_list[ii][jj] < self.diviation_ma_list[ii][jj-1]:
			sell_signal_3 = True

		if sell_signal_1 or sell_signal_2 or sell_signal_3:
			flag = True

		return flag

	def back_test(self,long_stand_deviation,short_stand_deviation):
		action_long_index_list =[]
		action_short_index_list =[]
		action_long_point_list =[]
		action_short_point_list = []
		action_datetime_list =[]
		action_revenue_list = []

		for ii in range(len(self.user_stock_code_list)):
			pos_flag = 0
			temp_action_long_index_list =[]
			temp_action_short_index_list =[]
			temp_action_long_point_list =[]
			temp_action_short_point_list = []
			temp_action_datetime_list =[]
			temp_action_revenue_list = []

			for jj in range(len(self.user_stock_price_data_list[ii])):
				if self.buy_signal(ii,jj,long_stand_deviation):
					temp_action_long_index_list.append(jj)
					temp_action_long_point_list.append(self.user_stock_uniform_price_list[ii][jj])
				if self.sell_signal(ii,jj,short_stand_deviation):
					temp_action_short_index_list.append(jj)
					temp_action_short_point_list.append(self.user_stock_uniform_price_list[ii][jj])
			'''
			for jj in range(len(self.user_stock_price_data_list[ii])):
				if self.buy_signal(ii,jj,long_stand_deviation):
					if pos_flag == 0:
						if len(temp_action_datetime_list) ==0:
							temp_action_revenue_list.append(1.0*self.trade_cost)
						else:
							temp_action_revenue_list.append(1.0*temp_action_revenue_list[-1]*self.trade_cost)
						temp_action_long_index_list.append(jj)
						temp_action_long_point_list.append(self.user_stock_uniform_price_list[ii][jj])
						temp_action_datetime_list.append(self.user_stock_datetime_list[ii][jj])
						pos_flag =1
					else:
						temp_action_revenue_list.append(temp_action_revenue_list[-1]*float(self.user_stock_uniform_price_list[ii][jj])/float(self.user_stock_uniform_price_list[ii][jj-1]))
				elif self.sell_signal(ii,jj,short_stand_deviation):
					if pos_flag ==1:
						temp_action_revenue_list.append(temp_action_revenue_list[-1]*float(self.user_stock_uniform_price_list[ii][jj])/float(self.user_stock_uniform_price_list[ii][jj-1])*self.trade_cost)
						temp_action_short_index_list.append(jj)
						temp_action_short_point_list.append(self.user_stock_uniform_price_list[ii][jj])
						temp_action_datetime_list.append(self.user_stock_datetime_list[ii][jj])
						pos_flag =0 
					else:
						if len(temp_action_datetime_list) ==0:
							temp_action_revenue_list.append(1.0)
						else:
							if self.short_flag ==1:
								temp_action_revenue_list.append(temp_action_revenue_list[-1]*(2-float(self.user_stock_uniform_price_list[ii][jj])/float(self.user_stock_uniform_price_list[ii][jj-1])))
							else:
								temp_action_revenue_list.append(temp_action_revenue_list[-1])
				else:
					if len(temp_action_datetime_list) ==0:
						temp_action_revenue_list.append(1.0)
					else:
						if pos_flag ==1:
							temp_action_revenue_list.append(temp_action_revenue_list[-1]*float(self.user_stock_uniform_price_list[ii][jj])/float(self.user_stock_uniform_price_list[ii][jj-1]))
						elif self.short_flag ==1:
							temp_action_revenue_list.append(temp_action_revenue_list[-1]*(2-float(self.user_stock_uniform_price_list[ii][jj])/float(self.user_stock_uniform_price_list[ii][jj-1])))
						else:
							temp_action_revenue_list.append(temp_action_revenue_list[-1])
			'''
			action_revenue_list.append(temp_action_revenue_list)
			action_long_index_list.append(temp_action_long_index_list)
			action_long_point_list.append(temp_action_long_point_list)
			action_short_index_list.append(temp_action_short_index_list)
			action_short_point_list.append(temp_action_short_point_list)
			action_datetime_list.append(temp_action_datetime_list)

		return [action_datetime_list,action_long_index_list,action_long_point_list,action_short_index_list,action_short_point_list,action_revenue_list]




	def eveluate_index(self,ma_days):
		print ma_days
		for ii in range(len(self.user_stock_code_list)):
			max_ =int(np.around(max(self.diviation_from_ma_list[ii])*100))
			min_ =int(np.around(min(self.diviation_from_ma_list[ii])*100))
			mean_ = int(np.around(100*np.mean(self.diviation_from_ma_list[ii])))
			median_ = int(np.around(100*np.median(self.diviation_from_ma_list[ii])))
			perc_95 =int(np.around(100*np.percentile(self.diviation_from_ma_list[ii],95)))
			perc_5 = int(np.around(100*np.percentile(self.diviation_from_ma_list[ii],5)))


			print self.user_stock_name_list[ii]+" &" + str(max_)+"\% &" +str(min_)+"\% &"+str(mean_)+"\% &"+str(median_) +"\% &" +str(perc_95)+"\% &" + str(perc_5)+"\% \\\ \hline"

	def calc_diviation_from_ma(self,ma_days):
		diviation_from_ma_list = []
		diviation_from_ma_std_list = []
		for ii in range(len(self.price_ma_list)):
			diviation_from_ma = []
			std = []
			for jj in range(len(self.price_ma_list[ii])):
				temp = float(self.user_stock_price_data_list[ii][jj])/float(self.price_ma_list[ii][jj])-1
				diviation_from_ma.append(temp)

				temp_std = np.std(diviation_from_ma[max(jj-ma_days+1,0):jj+1])
				std.append(temp_std)

			diviation_from_ma_list.append(diviation_from_ma)
			diviation_from_ma_std_list.append(std)
		return diviation_from_ma_list,diviation_from_ma_std_list


		



	def plot(self,ma_days):
		for ii in range(len(self.user_stock_code_list)):
			plt.figure(ii)
			ax = plt.gca()
			plt.plot(self.diviation_from_ma_list[ii],color = "b",linewidth = 2.5,label = "Price Diviation From MA")

			std =[]

			for jj in range(len(self.diviation_from_ma_list[ii])):
				temp_std = np.std(self.diviation_from_ma_list[ii][max(jj-ma_days+1,0):jj+1])
				std.append(temp_std)

			plt.plot(std,color = "r",linewidth = 2.5,label= "Stand Diviation")
			plt.xlabel("Time /day")
			plt.ylabel("Diviation Percentage")
			plt.legend(loc="upper right")

			xticks = ax.get_xticks()
			offset = xticks[1]-xticks[0]
			date_time = []
			for jj in range(len(xticks)):
				if jj *int(offset)>= len(self.user_stock_datetime_list[ii]):
					break
				date_time.append(self.user_stock_datetime_list[ii][jj*int(offset)])
			ax.set_xticklabels(date_time,rotation = 30)

			plt.title(self.user_stock_name_list[ii]+" Price Diviation From "+str(ma_days)+" MA")
			plt.grid(True)
			fig = plt.gcf()
			fig.set_size_inches(24, 15)
			fig.savefig("../input/DeviationFromMA/png_"+str(ma_days)+"/"+self.user_stock_code_list[ii]+".png",dpi = 100)
			plt.close(fig)

	def load_excel(self,file_path):
		user_stock_code_list = []
		user_stock_name_list = []

		wb = load_workbook(file_path)
		sheet_names = wb.get_sheet_names()
		ws = wb.get_sheet_by_name(sheet_names[0])

		row_num = len(ws.rows)
		for ii in range(1,row_num):
			user_stock_code_list.append(str(ws.cell(row=ii+1,column = 1).value))
			user_stock_name_list.append(ws.cell(row=ii+1,column = 2).value)

		return user_stock_code_list,user_stock_name_list


if __name__ == '__main__':
	days = [60,120,200,250]
	for day in days:
		case = DeviationFromMA(day)
